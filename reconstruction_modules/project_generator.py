# ======================================================
# reconstruction_modules/project_generator.py
# Project Generation and Excel Template Creation
# ======================================================

from config import *
import geopandas as gpd
import pandas as pd
import re
from datetime import datetime
from pyproj import Transformer
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ======================================================
# Project Generation Logic
# ======================================================

def _to_wgs84_lat_lon(point, source_crs):
    """Convert a projected point to WGS84 latitude/longitude."""
    if point is None:
        return None, None
    try:
        if source_crs is None:
            source_crs = TARGET_CRS
        transformer = Transformer.from_crs(source_crs, "EPSG:4326", always_xy=True)
        lon, lat = transformer.transform(point.x, point.y)
        return round(lat, 6), round(lon, 6)
    except Exception:
        return None, None


def _is_arabic_column_name(name):
    """Return True if a column name contains Arabic characters."""
    if name is None:
        return False
    return re.search(r"[\u0600-\u06FF]", str(name)) is not None


def enrich_projects_with_reference_points(projects_df, infrastructure_layers, max_distance_m=3000):
    """Attach nearest named hospital/clinic/street to each project using local GIS layers."""
    if projects_df is None or projects_df.empty:
        return projects_df
    if "Latitude" not in projects_df.columns or "Longitude" not in projects_df.columns:
        return projects_df
    if infrastructure_layers is None:
        return projects_df

    export_df = projects_df.copy()
    valid_coords = export_df["Latitude"].notna() & export_df["Longitude"].notna()
    if not valid_coords.any():
        return export_df

    candidates = []
    source_specs = [
        ("hospitals", "Hospital/Clinic", "مستشفى/عيادة"),
        ("streets", "Street", "شارع"),
    ]

    for layer_key, type_en, type_ar in source_specs:
        layer = infrastructure_layers.get(layer_key)
        if layer is None or layer.empty:
            continue

        layer = layer.copy()
        if layer.crs is None:
            layer = layer.set_crs(TARGET_CRS, allow_override=True)
        elif str(layer.crs) != TARGET_CRS:
            layer = layer.to_crs(TARGET_CRS)

        layer["geometry"] = layer.geometry.representative_point()
        name_col = None
        for col in ["name", "Name", "NAME", "name_en", "name_ar", "ref", "RoadName", "road_name"]:
            if col in layer.columns:
                name_col = col
                break

        if name_col is not None:
            names = layer[name_col].astype(str).str.strip()
        elif layer_key == "streets" and "RoadType" in layer.columns:
            names = layer["RoadType"].astype(str).str.strip() + " Road"
        else:
            names = pd.Series([f"{type_en} point"] * len(layer), index=layer.index, dtype="object")

        layer["ref_name"] = names.replace({"": f"{type_en} point", "nan": f"{type_en} point", "None": f"{type_en} point"})
        layer["ref_type_en"] = type_en
        layer["ref_type_ar"] = type_ar
        candidates.append(layer[["geometry", "ref_name", "ref_type_en", "ref_type_ar"]])

    if not candidates:
        return export_df

    refs = gpd.GeoDataFrame(pd.concat(candidates, ignore_index=True), geometry="geometry", crs=TARGET_CRS)
    project_points = gpd.GeoDataFrame(
        export_df[valid_coords].copy(),
        geometry=gpd.points_from_xy(export_df.loc[valid_coords, "Longitude"], export_df.loc[valid_coords, "Latitude"]),
        crs="EPSG:4326",
    ).to_crs(TARGET_CRS)

    nearest = gpd.sjoin_nearest(
        project_points[["geometry"]],
        refs,
        how="left",
        distance_col="distance_to_reference_m",
        max_distance=max_distance_m,
    )
    # Handle duplicate matches (same project matched to multiple equidistant refs).
    nearest = nearest.sort_values("distance_to_reference_m", ascending=True, na_position="last")
    nearest = nearest[~nearest.index.duplicated(keep="first")]

    default_names = {"", "N/A", "None", "Unknown", "Hospital/Clinic", "Street/Road Segment"}
    for idx, nearest_row in nearest.iterrows():
        candidate_name = nearest_row.get("ref_name")
        candidate_type = nearest_row.get("ref_type_en")

        if candidate_name is None or pd.isna(candidate_name):
            continue

        existing_name = str(export_df.at[idx, "Reference_Point_Name"]) if "Reference_Point_Name" in export_df.columns else "N/A"
        if existing_name in default_names:
            export_df.at[idx, "Reference_Point_Name"] = str(candidate_name)
        if "Reference_Point_Type" not in export_df.columns or str(export_df.at[idx, "Reference_Point_Type"]) in default_names:
            export_df.at[idx, "Reference_Point_Type"] = str(candidate_type or "Reference Point")

        val = nearest_row.get("distance_to_reference_m")
        export_df.at[idx, "Distance_To_Reference_m"] = round(float(val), 1) if pd.notna(val) else None

    return export_df


def generate_projects_from_zones(hex_gdf, selected_projects, top_n=20):
    """Generate reconstruction projects from prioritized zones for selected project types"""
    print(f"   - Generating projects for selected types: {', '.join(selected_projects)}...")

    projects = []
    project_counter = 1

    # Generate 20-30 projects per selected type
    projects_per_type = 25

    for project_type in selected_projects:
        print(f"   - Generating {projects_per_type} projects for {project_type}...")
        type_projects = generate_projects_by_type(hex_gdf, project_type, projects_per_type, project_counter)
        projects.extend(type_projects)
        project_counter += len(type_projects)

    projects_df = pd.DataFrame(projects)
    
    # Add final priority rank
    if not projects_df.empty:
        projects_df['Final_Priority_Rank'] = range(1, len(projects_df) + 1)

    print(f"   - Generated {len(projects)} reconstruction projects")

    return projects_df

def generate_projects_by_type(hex_gdf, project_type, count, start_id):
    """Generate projects for a specific infrastructure type"""
    projects = []
    
    # Normalize project type name
    project_type_lower = project_type.lower().replace('_', ' ')
    
    # Filter zones based on project type
    if 'hospital' in project_type_lower or 'healthcare' in project_type_lower:
        if 'major' in project_type_lower:
            relevant_zones = hex_gdf[hex_gdf['hospitals_count'] > 0].head(count)
            infra_type = 'Major Hospital Reconstruction'
            cost_multiplier = 3.0
        else:
            relevant_zones = hex_gdf[hex_gdf['hospitals_count'] > 0].head(count)
            infra_type = 'Healthcare Facilities'
            cost_multiplier = 2.0
    elif 'education' in project_type_lower or 'school' in project_type_lower:
        relevant_zones = hex_gdf[hex_gdf['schools_count'] > 0].head(count)
        infra_type = 'Education Facilities'
        cost_multiplier = 1.5
    elif 'universit' in project_type_lower:
        relevant_zones = hex_gdf[hex_gdf['universities_count'] > 0].head(count)
        infra_type = 'University Facilities'
        cost_multiplier = 3.0
    elif 'transport' in project_type_lower or 'street' in project_type_lower or 'road' in project_type_lower:
        relevant_zones = hex_gdf[hex_gdf['streets_count'] > 0].head(count)
        infra_type = 'Transportation & Streets'
        cost_multiplier = 1.0
    elif 'municipal' in project_type_lower:
        relevant_zones = hex_gdf[hex_gdf['municipalities_count'] > 0].head(count)
        infra_type = 'Municipal Services'
        cost_multiplier = 1.8
    elif 'utilit' in project_type_lower or 'water' in project_type_lower or 'fuel' in project_type_lower:
        relevant_zones = hex_gdf[(hex_gdf['water_util_count'] > 0) | (hex_gdf['fuel_util_count'] > 0)].head(count)
        infra_type = 'Utilities Infrastructure'
        cost_multiplier = 2.5
    else:
        relevant_zones = hex_gdf.head(count)
        infra_type = 'General Reconstruction'
        cost_multiplier = 1.0

    # If not enough zones with that infrastructure, use top priority zones
    if len(relevant_zones) < count:
        additional_needed = count - len(relevant_zones)
        additional_zones = hex_gdf[~hex_gdf.index.isin(relevant_zones.index)].head(additional_needed)
        relevant_zones = pd.concat([relevant_zones, additional_zones])

    for idx, (_, row) in enumerate(relevant_zones.iterrows()):
        damage_level = row.get('damage_count', 0)
        required_units = max(int(damage_level * UNITS_PER_DAMAGE_SITE), MINIMUM_UNITS)
        
        strategy = row.get('rebuilding_strategy', {}).get('strategy', 'Balanced_Reconstruction')
        if strategy == 'Basic_Infrastructure_First':
            required_units = max(required_units, MINIMUM_UNITS_HIGH_DAMAGE)
        elif strategy == 'Street_Reconstruction_Priority':
            required_units = min(required_units, MAXIMUM_UNITS_LOW_DAMAGE)

        cost_per_unit = COST_ESTIMATES.get('per_unit', 50000) * cost_multiplier
        timeline_months = TIMELINE_ESTIMATES.get(strategy, 12)
        
        # Get population data
        population = row.get('population_total', 0)
        pop_density = row.get('population_density', 0)
        pop_prewar = row.get('population_prewar', 0)
        is_idp_hotspot = row.get('is_idp_hotspot', False)
        
        # Calculate displacement
        if pop_prewar > 0:
            displacement_rate = ((pop_prewar - population) / pop_prewar) * 100
        else:
            displacement_rate = 0
        
        # Determine urgency level based on density
        if pop_density > 30000:
            urgency = 'CRITICAL - Extreme Overcrowding'
        elif pop_density > 15000:
            urgency = 'URGENT - High Density'
        elif pop_density > 8000:
            urgency = 'HIGH - Above Normal'
        else:
            urgency = 'NORMAL'

        # Get zone centroid coordinates
        zone_centroid = row.geometry.centroid
        lat, lon = _to_wgs84_lat_lon(zone_centroid, hex_gdf.crs)

        # Identify the dominant reference point for this project
        reference_point_type = "General Zone"
        reference_point_name = "N/A"
        if infra_type == 'Major Hospital Reconstruction':
            reference_point_type = "Major Hospital"
            reference_point_name = row.get('major_hospital_names', 'N/A')
        elif infra_type == 'Healthcare Facilities':
            if row.get('major_hospitals_count', 0) > 0:
                reference_point_type = "Major Hospital"
                reference_point_name = row.get('major_hospital_names', 'N/A')
            else:
                reference_point_type = "Hospital/Clinic"
        elif infra_type == 'Education Facilities':
            reference_point_type = "School"
        elif infra_type == 'University Facilities':
            reference_point_type = "University"
        elif infra_type == 'Utilities Infrastructure':
            water_count = row.get('water_util_count', 0)
            fuel_count = row.get('fuel_util_count', 0)
            if water_count >= fuel_count and water_count > 0:
                reference_point_type = "Water Utility"
            elif fuel_count > 0:
                reference_point_type = "Energy/Fuel Utility"
            else:
                reference_point_type = "Utility Facility"
        elif infra_type == 'Municipal Services':
            reference_point_type = "Municipal Facility"
        elif infra_type == 'Transportation & Streets':
            reference_point_type = "Street/Road Segment"
        
        project = {
            'Project_ID': f"GAZA-{start_id + idx:03d}",
            'Project_Name': f"{infra_type} - {row.get('primary_municipality', 'Gaza')} (Lat: {lat}, Lon: {lon})",
            'Zone_ID': row.get('zone_id', f"ZONE-{idx+1:03d}"),
            'Latitude': lat,
            'Longitude': lon,
            'Municipality': row.get('primary_municipality', 'Unknown'),
            'Priority_Rank': start_id + idx,
            'AI_Score': round(row.get('ai_score', 0), 3),
            'Damage_Level': damage_level,
            'Population_Current': int(population),
            'Population_PreWar': int(pop_prewar),
            'Displacement_Rate': round(displacement_rate, 1),
            'Population_Density': round(pop_density, 0),
            'IDP_Hotspot': 'YES' if is_idp_hotspot else 'NO',
            'Urgency_Level': urgency,
            'Rebuilding_Strategy': strategy,
            'Required_Units': required_units,
            'Timeline_Months': timeline_months,
            'Infrastructure_Type': infra_type,
            'Reference_Point_Type': reference_point_type,
            'Reference_Point_Name': reference_point_name,
            'Project_Type': project_type,
            'Status': 'Planned',
            'Funding_Source': 'TBD',
            'Implementing_Agency': 'TBD',
            'Beneficiaries_Current': int(population * 0.8) if population > 0 else 0,
            'Beneficiaries_PostReturn': int(pop_prewar * 0.8) if pop_prewar > 0 else 0,
            'Expert_Notes': row.get('expert_explanation', '')
        }
        projects.append(project)

    return projects

def generate_street_reconstruction_projects(streets_gdf, top_n=100):
    """Generate street reconstruction projects with priorities and phases"""
    print(f"   - Generating street reconstruction projects...")
    
    if streets_gdf.empty:
        return pd.DataFrame()
    
    # Filter damaged streets only
    damaged_streets = streets_gdf[streets_gdf.get('damage_severity', 0) > 0].copy()
    
    if damaged_streets.empty:
        print("   - No damaged streets found")
        return pd.DataFrame()
    
    # Sort by reconstruction priority
    damaged_streets = damaged_streets.sort_values('reconstruction_priority', ascending=False)
    
    # Take top N streets
    top_streets = damaged_streets.head(top_n)
    
    projects = []
    for idx, street in top_streets.iterrows():
        damage = street.get('damage_severity', 0)
        road_type = street.get('RoadType', 'Unknown')
        municipality = street.get('municipality', 'Unknown')
        length_km = street.get('Length_km', 0)
        priority_score = street.get('reconstruction_priority', 0)
        
        # Calculate timeline (2 months per km * damage multiplier)
        damage_multiplier = max(0.5, float(damage) / 3.0)
        reconstruction_months = max(1, int(length_km * 2 * damage_multiplier))
        
        # Determine priority level
        if priority_score >= 30:
            priority_level = 'Critical'
            phase = 'Emergency (0-12 months)'
        elif priority_score >= 20:
            priority_level = 'High'
            phase = 'Basics (12-30 months)'
        elif priority_score >= 10:
            priority_level = 'Medium'
            phase = 'Development (30-54 months)'
        else:
            priority_level = 'Low'
            phase = 'Improvement (54-84 months)'
        
        # Get coordinates
        centroid = street.geometry.centroid
        lat, lon = _to_wgs84_lat_lon(centroid, streets_gdf.crs)
        
        project = {
            'Project_ID': f"STREET-{idx+1:04d}",
            'Project_Name': f"Street Reconstruction - {road_type} - {municipality} (Lat: {lat}, Lon: {lon})",
            'Latitude': lat,
            'Longitude': lon,
            'Municipality': municipality,
            'Road_Type': road_type,
            'Reference_Point_Type': 'Street/Road Segment',
            'Reference_Point_Name': f"{road_type} Road",
            'Length_km': round(length_km, 2),
            'Damage_Severity': damage,
            'Priority_Level': priority_level,
            'Priority_Score': round(priority_score, 1),
            'Reconstruction_Phase': phase,
            'Timeline_Months': reconstruction_months,
            'Infrastructure_Type': 'Transportation & Streets',
            'Status': 'Planned',
            'Near_Hospital': 'YES' if street.get('near_hospital', False) else 'NO',
            'Major_Artery': 'YES' if street.get('is_major_artery', False) else 'NO'
        }
        projects.append(project)
    
    projects_df = pd.DataFrame(projects)
    
    if not projects_df.empty:
        total_length = projects_df['Length_km'].sum()
        print(f"   - Generated {len(projects_df)} street reconstruction projects")
        print(f"   - Total length: {total_length:.1f} km")
    
    return projects_df

# ======================================================
# Excel Export Functionality
# ======================================================

def export_projects_to_excel(projects_df, output_path=None):
    """Export projects to Excel with formatting"""
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f"gaza_reconstruction_projects_{timestamp}.xlsx"

    print(f"   - Exporting projects to {output_path}...")

    try:
        # Create a copy for export
        export_df = projects_df.copy()

        # Normalize coordinate column names for export consistency.
        if "Latitude" not in export_df.columns and "lat" in export_df.columns:
            export_df["Latitude"] = export_df["lat"]
        if "Longitude" not in export_df.columns and "lon" in export_df.columns:
            export_df["Longitude"] = export_df["lon"]

        area_sources = ["Municipality", "primary_municipality", "Governorate", "municipality"]
        if "Area_Name" not in export_df.columns:
            area_name = pd.Series("Unknown", index=export_df.index, dtype="object")
            for col in area_sources:
                if col in export_df.columns:
                    values = export_df[col].astype(str).str.strip()
                    valid = values.ne("") & values.ne("nan") & values.ne("None") & values.ne("Unknown")
                    area_name = area_name.where(~valid, values)
            export_df["Area_Name"] = area_name

        if "Google_Maps_Link" not in export_df.columns and {"Latitude", "Longitude"}.issubset(export_df.columns):
            export_df["Google_Maps_Link"] = export_df.apply(
                lambda row: f"https://www.google.com/maps?q={row['Latitude']},{row['Longitude']}"
                if pd.notna(row["Latitude"]) and pd.notna(row["Longitude"]) else "",
                axis=1
            )

        export_df = _drop_redundant_export_columns(export_df)

        preferred_order = [
            "Project_ID",
            "Project_Name",
            "Reference_Point_Type",
            "Reference_Point_Name",
            "Latitude",
            "Longitude",
            "Google_Maps_Link",
            "Area_Name",
            "Municipality",
            "Zone_ID",
        ]
        existing_preferred = [col for col in preferred_order if col in export_df.columns]
        remaining_cols = [col for col in export_df.columns if col not in existing_preferred]
        export_df = export_df[existing_preferred + remaining_cols]

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main projects sheet
            export_df.to_excel(writer, sheet_name='Reconstruction_Projects', index=False)

            # Summary sheet
            summary_data = create_project_summary(projects_df)
            summary_data.to_excel(writer, sheet_name='Project_Summary', index=False)

            # Strategy breakdown sheet
            strategy_breakdown = create_strategy_breakdown(projects_df)
            strategy_breakdown.to_excel(writer, sheet_name='Strategy_Breakdown', index=False)
            
            # Population density by municipality
            if 'Population_Current' in projects_df.columns:
                pop_by_mun = projects_df.groupby('Municipality').agg({
                    'Population_Current': 'sum',
                    'Population_PreWar': 'sum',
                    'Population_Density': 'mean',
                    'Beneficiaries_Current': 'sum',
                    'Project_ID': 'count'
                }).round(0)
                pop_by_mun.columns = ['Current_Population', 'PreWar_Population', 'Avg_Density', 'Total_Beneficiaries', 'Number_of_Projects']
                pop_by_mun.to_excel(writer, sheet_name='Population_by_Municipality')

            _style_workbook(writer)

        print(f"   - Successfully exported {len(projects_df)} projects to Excel")
        return output_path

    except Exception as e:
        print(f"   - Error exporting to Excel: {e}")
        return None

# ======================================================
# Summary and Analytics
# ======================================================

def create_project_summary(projects_df):
    """Create summary statistics for projects"""
    required_units = (
        pd.to_numeric(projects_df['Required_Units'], errors='coerce').fillna(0)
        if 'Required_Units' in projects_df.columns
        else pd.Series(0, index=projects_df.index, dtype='float64')
    )
    timeline_months = (
        pd.to_numeric(projects_df['Timeline_Months'], errors='coerce').fillna(0)
        if 'Timeline_Months' in projects_df.columns
        else pd.Series(0, index=projects_df.index, dtype='float64')
    )
    priority_rank = (
        pd.to_numeric(projects_df['Priority_Rank'], errors='coerce').fillna(999999)
        if 'Priority_Rank' in projects_df.columns
        else pd.Series(999999, index=projects_df.index, dtype='float64')
    )
    strategy_col = (
        projects_df['Rebuilding_Strategy'].astype(str)
        if 'Rebuilding_Strategy' in projects_df.columns
        else pd.Series('', index=projects_df.index, dtype='object')
    )

    # Population statistics
    total_current_pop = projects_df['Population_Current'].sum() if 'Population_Current' in projects_df.columns else 0
    total_prewar_pop = projects_df['Population_PreWar'].sum() if 'Population_PreWar' in projects_df.columns else 0
    avg_density = projects_df['Population_Density'].mean() if 'Population_Density' in projects_df.columns else 0
    
    summary = {
        'Total_Projects': len(projects_df),
        'Total_Required_Units': int(required_units.sum()),
        'Average_Timeline_Months': float(timeline_months.mean()) if len(timeline_months) > 0 else 0,
        'Total_Current_Population': int(total_current_pop),
        'Total_PreWar_Population': int(total_prewar_pop),
        'Average_Population_Density': round(avg_density, 0),
        'Total_Beneficiaries_Current': projects_df['Beneficiaries_Current'].sum() if 'Beneficiaries_Current' in projects_df.columns else 0,
        'Total_Beneficiaries_PostReturn': projects_df['Beneficiaries_PostReturn'].sum() if 'Beneficiaries_PostReturn' in projects_df.columns else 0,
        'High_Priority_Projects': int((priority_rank <= 5).sum()),
        'Critical_Urgency_Projects': len(projects_df[projects_df['Urgency_Level'] == 'CRITICAL - Extreme Overcrowding']) if 'Urgency_Level' in projects_df.columns else 0,
        'IDP_Hotspot_Projects': len(projects_df[projects_df['IDP_Hotspot'] == 'YES']) if 'IDP_Hotspot' in projects_df.columns else 0,
        'Basic_Infrastructure_First': int((strategy_col == 'Basic_Infrastructure_First').sum()),
        'Balanced_Reconstruction': int((strategy_col == 'Balanced_Reconstruction').sum()),
        'Street_Reconstruction_Priority': int((strategy_col == 'Street_Reconstruction_Priority').sum())
    }

    return pd.DataFrame([summary])

def create_strategy_breakdown(projects_df):
    """Create breakdown by rebuilding strategy"""
    if projects_df is None or projects_df.empty:
        return pd.DataFrame(columns=['Strategy', 'Project_Count', 'Total_Units', 'Avg_Timeline'])

    strategy_series = (
        projects_df['Rebuilding_Strategy'].astype(str)
        if 'Rebuilding_Strategy' in projects_df.columns
        else pd.Series(['Not_Specified'] * len(projects_df), index=projects_df.index, dtype='object')
    )
    work_df = projects_df.copy()
    work_df['Rebuilding_Strategy'] = strategy_series.replace({'': 'Not_Specified', 'nan': 'Not_Specified', 'None': 'Not_Specified'})
    if 'Required_Units' not in work_df.columns:
        work_df['Required_Units'] = 0
    if 'Timeline_Months' not in work_df.columns:
        work_df['Timeline_Months'] = 0

    strategy_breakdown = work_df.groupby('Rebuilding_Strategy').agg({
        'Project_ID': 'count',
        'Required_Units': 'sum',
        'Timeline_Months': 'mean'
    }).reset_index()

    strategy_breakdown.columns = ['Strategy', 'Project_Count', 'Total_Units', 'Avg_Timeline']

    return strategy_breakdown


def _style_workbook(writer):
    """Apply a professional baseline style to all Excel sheets."""
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="center")

    for sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.auto_filter.ref = ws.dimensions

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
                if cell.row > 1:
                    cell.alignment = body_alignment
            adjusted = min(max(12, max_len + 2), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        headers = {cell.value: cell.column for cell in ws[1]}
        for coord_col in ["Latitude", "Longitude"]:
            if coord_col in headers:
                col_letter = get_column_letter(headers[coord_col])
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row}"].number_format = "0.000000"


def _drop_redundant_export_columns(df):
    """Drop columns that are duplicated by name or fully duplicated by content."""
    if df is None or df.empty:
        return df

    out = df.copy()

    # Explicitly remove internal/legacy columns not needed in final templates.
    drop_if_exists = ["Reference_Point_Type_AR", "Estimated_Cost"]
    out = out.drop(columns=[c for c in drop_if_exists if c in out.columns], errors="ignore")

    # 1) Remove exact duplicate column names (keep first occurrence).
    out = out.loc[:, ~out.columns.duplicated(keep="first")]

    # 2) Remove Arabic-named columns (English-only headers in exports).
    arabic_cols = [col for col in out.columns if _is_arabic_column_name(col)]
    if arabic_cols:
        out = out.drop(columns=arabic_cols)

    def _normalized(series):
        return series.astype("string").fillna("<NA>").str.strip()

    # 3) Remove any remaining columns with identical full content (keep first).
    kept = []
    for col in list(out.columns):
        same_as_kept = False
        for kcol in kept:
            if _normalized(out[col]).equals(_normalized(out[kcol])):
                same_as_kept = True
                break
        if not same_as_kept:
            kept.append(col)

    return out[kept]
